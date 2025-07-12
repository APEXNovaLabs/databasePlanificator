import pandas as pd
import datetime
import asyncio
import aiomysql
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from Contrat.fonctionnalités.connexionDB import DBConnection


# --- Fonction de récupération des données de facture pour un client spécifique ---
async def obtenirDataFactureClient(pool, client_id: int, year: int, month: int):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT cl.nom            AS client_nom,
                           COALESCE(cl.prenom, '') AS client_prenom, /* Correction: Gérer les prénoms NULL */
                           cl.adresse        AS client_adresse,
                           cl.telephone      AS client_telephone,
                           cl.categorie      AS client_categorie,
                           f.date_traitement AS `Date de traitement`,
                           tt.typeTraitement AS `Traitement (Type)`,
                           pd.statut         AS `Etat traitement`,
                           f.etat            AS `Etat paiement (Payée ou non)`,
                           -- Utilise le nouveau_montant de Historique_prix si disponible (le plus récent),
                           -- sinon utilise le montant original de Facture.
                           COALESCE(
                                   (SELECT hp.new_amount
                                    FROM Historique_prix hp
                                    WHERE hp.facture_id = f.facture_id
                                    ORDER BY hp.change_date DESC, hp.history_id
                                                            DESC -- Prend le plus récent, history_id pour briser l'égalité
                                    LIMIT 1),
                                   f.montant
                           )                 AS montant_facture
                    FROM Facture f
                             JOIN
                         PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                             JOIN
                         Planning p ON pd.planning_id = p.planning_id
                             JOIN
                         Traitement tr ON p.traitement_id = tr.traitement_id
                             JOIN
                         TypeTraitement tt ON tr.id_type_traitement = tt.id_type_traitement
                             JOIN
                         Contrat co ON tr.contrat_id = co.contrat_id
                             JOIN
                         Client cl ON co.client_id = cl.client_id
                    WHERE cl.client_id = %s
                      AND YEAR(f.date_traitement) = %s
                      AND MONTH(f.date_traitement) = %s
                    ORDER BY f.date_traitement;
                    """
            await cursor.execute(query, (client_id, year, month))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des données de facture : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction utilitaire pour obtenir le client_id et nom complet de tous les clients ---
async def obtenirTousClient(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id, CONCAT(nom, ' ', COALESCE(prenom, '')) AS full_name /* Correction: Gérer les prénoms NULL */
                    FROM Client
                    ORDER BY full_name;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération de la liste des clients : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction utilitaire pour obtenir le client_id à partir du nom/prénom ---
async def obtenirIDClientAvecNom(pool, client_name: str):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id
                    FROM Client
                    WHERE CONCAT(nom, ' ', COALESCE(prenom, '')) = %s /* Correction: Gérer les prénoms NULL */
                    LIMIT 1;
                    """
            await cursor.execute(query, (client_name,))
            result = await cursor.fetchone()
            return result['client_id'] if result else None
    except Exception as e:
        print(f"Erreur lors de la recherche du client par nom : {e}")
        return None
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour récupérer la liste des clients, le nombre de factures et les mois des factures ---
async def obtenirInformationsClients(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT c.client_id,
                           CONCAT(c.nom, ' ', COALESCE(c.prenom, '')) AS nomComplet, /* Correction: Gérer les prénoms NULL */
                           COUNT(f.facture_id)          AS total_factures,
                           GROUP_CONCAT(DISTINCT
                                        CONCAT(YEAR(f.date_traitement), '-', LPAD(MONTH(f.date_traitement), 2, '0'))
                                        ORDER BY YEAR(f.date_traitement) DESC, MONTH(f.date_traitement) DESC SEPARATOR
                                        ', ')           AS facture_months
                    FROM Client c
                             LEFT JOIN Contrat co ON c.client_id = co.client_id
                             LEFT JOIN Traitement tr ON co.contrat_id = tr.contrat_id
                             LEFT JOIN Planning p ON tr.traitement_id = p.traitement_id
                             LEFT JOIN PlanningDetails pd ON p.planning_id = pd.planning_id
                             LEFT JOIN Facture f ON pd.planning_detail_id = f.planning_detail_id
                    GROUP BY c.client_id, nomComplet
                    ORDER BY nomComplet;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération du nombre de factures par client et des mois : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour générer le fichier Excel de la facture client ---
def genererFactureExcel(data: list[dict], client_full_name: str, year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()

    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"{safe_client_name}-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facture {client_full_name} {month_name_fr}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    ligneActuelle = 1

    # Informations du client (en-tête)
    if data:
        infoClient = data[0]
        # client_prenom est maintenant garanti d'être une chaîne vide ou non vide grâce à COALESCE
        affichageNomClient = f"{infoClient['client_nom']} {infoClient['client_prenom']}"
        if infoClient['client_categorie'] != 'Particulier':
            affichageNomClient = f"{infoClient['client_nom']} (Responsable: {infoClient['client_prenom'] if infoClient['client_prenom'] else 'N/A'})"

        ws.cell(row=ligneActuelle, column=1, value="Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=affichageNomClient)
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Adresse :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_adresse'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_telephone'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_categorie'])
        ligneActuelle += 1

    ligneActuelle += 1  # Ligne vide

    # Tableau des traitements
    table_headers = ['Date de traitement', 'Traitement (Type)', 'Etat traitement', 'Etat paiement (Payée ou non)',
                     'Montant']
    num_table_cols = len(table_headers) # Dynamiser le nombre de colonnes

    # Ligne "Facture du mois de:"
    ws.cell(row=ligneActuelle, column=1, value=f"Facture du mois de : {month_name_fr} {year}").font = header_font
    ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=num_table_cols)
    ws.cell(row=ligneActuelle, column=1).alignment = Alignment(horizontal='center')
    ligneActuelle += 2  # Deux lignes vides après pour la séparation avec le tableau


    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=ligneActuelle, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    ligneActuelle += 1

    if not data:
        ws.cell(row=ligneActuelle, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour ce mois.").border = thin_border
        ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=len(table_headers))
        ligneActuelle += 1
    else:
        df_invoice_data = pd.DataFrame(data)
        # Sélectionner et réordonner les colonnes pour l'affichage du tableau, incluant le nouveau montant_facture
        df_display = df_invoice_data[
            ['Date de traitement', 'Traitement (Type)', 'Etat traitement', 'Etat paiement (Payée ou non)',
             'montant_facture']]
        df_display.rename(columns={'montant_facture': 'Montant'}, inplace=True)  # Rename for display

        # Écrire les données du tableau et appliquer la couleur
        for r_idx, row_data in enumerate(df_display.values.tolist(), start=ligneActuelle):
            # Déterminer la couleur de remplissage basée sur l'état de paiement
            payment_status = row_data[3]  # 'Etat paiement (Payée ou non)' est la 4ème colonne (index 3)
            fill_to_apply = None
            if payment_status == 'Payé':
                fill_to_apply = green_fill
            elif payment_status == 'Non payé':
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            ligneActuelle += 1

    ligneActuelle += 1  # Ligne vide avant les totaux

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)  # Utilise le DataFrame complet avec montant_facture

        # Total par type de traitement (seulement payé)
        total_by_type_paid = df_calc[df_calc['Etat paiement (Payée ou non)'] == 'Payé'].groupby('Traitement (Type)')[
            'montant_facture'].sum()

        if not total_by_type_paid.empty:
            ws.cell(row=ligneActuelle, column=1, value="Facture total pour :").font = bold_font
            ligneActuelle += 1
            for service_type, total_amount in total_by_type_paid.items():
                ws.cell(row=ligneActuelle, column=2, value=f"{service_type} (Payé)").font = bold_font
                ws.cell(row=ligneActuelle, column=3, value=total_amount).font = bold_font
                ligneActuelle += 1
        else:
            ws.cell(row=ligneActuelle, column=1,
                    value="Aucun montant payé pour les types de traitement ce mois.").font = bold_font
            ligneActuelle += 1

        ligneActuelle += 1  # Ligne vide

        # Montant total pour tous les traitements effectués (même non payés)
        grand_total = df_calc['montant_facture'].sum()
        ws.cell(row=ligneActuelle, column=1, value="Montant total des traitements effectués ce mois :").font = bold_font
        ws.cell(row=ligneActuelle, column=3, value=grand_total).font = bold_font
        ligneActuelle += 1

    # Ajuster la largeur des colonnes
    # Utiliser le nombre de colonnes du tableau pour l'ajustement
    max_col_for_width = len(table_headers)

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                # openpyxl gère les objets datetime.date et datetime.datetime nativement,
                # donc str() est suffisant pour obtenir une longueur de chaîne.
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")


# --- Fonction principale pour exécuter la génération de facture ---
async def generationFactureClient():
    pool = None
    try:
        pool = await DBConnection()
        if pool is None:
            print("Échec de la connexion à la base de données. Annulation de l'opération.")
            return

        # --- Affichage du tableau des clients et de leurs nombres de factures ---
        print("\n--- Aperçu des clients et de leurs factures ---")
        informationsClient = await obtenirInformationsClients(pool)

        if informationsClient:
            # Créer un DataFrame pour un affichage tabulaire clair
            df_counts = pd.DataFrame(informationsClient)
            df_counts.rename(columns={'nomComplet': 'Nom du Client', 'total_factures': 'Nb. Factures',
                                      'facture_months': 'Mois des Factures'}, inplace=True)
            df_counts.index = df_counts.index + 1  # Pour commencer l'index à 1
            print(df_counts[['Nom du Client', 'Nb. Factures', 'Mois des Factures']].to_string())
            print("\n")  # Ligne vide pour la clarté
        else:
            print("Aucun client trouvé ou aucune facture enregistrée.")

        # --- Sélection du client ---
        clients_for_selection = await obtenirTousClient(pool)  # Réutiliser get_all_clients pour la sélection

        client_map = {}
        for i, client in enumerate(clients_for_selection):
            client_map[str(i + 1)] = client  # Store client_id for selection by number

        client_id_for_invoice = None
        client_full_name_for_invoice = None

        while client_id_for_invoice is None:
            choice = input(
                "Veuillez entrer le numéro du client dans la liste ci-dessus, ou son nom complet (Nom Prénom) : ").strip()

            if choice.isdigit():
                if choice in client_map:
                    selected_client = client_map[choice]
                    client_id_for_invoice = selected_client['client_id']
                    client_full_name_for_invoice = selected_client['full_name']
                    print(f"Client sélectionné : {client_full_name_for_invoice}")
                else:
                    print("Numéro invalide. Veuillez réessayer.")
            else:
                client_full_name_for_invoice = choice
                client_id_for_invoice = await obtenirIDClientAvecNom(pool, client_full_name_for_invoice)
                if client_id_for_invoice is None:
                    print(f"Client '{client_full_name_for_invoice}' non trouvé. Veuillez vérifier le nom et réessayer.")
                else:
                    print(f"Client trouvé : {client_full_name_for_invoice}")

        # --- Sélection du mois et de l'année pour la facture ---
        try:
            async with pool.acquire() as conn_for_months:
                async with conn_for_months.cursor(aiomysql.DictCursor) as cursor:
                    query_client_months = """
                                          SELECT DISTINCT YEAR(f.date_traitement)  AS annee,
                                                          MONTH(f.date_traitement) AS mois
                                          FROM Facture f
                                                   JOIN PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                                                   JOIN Planning p ON pd.planning_id = p.planning_id
                                                   JOIN Traitement tr ON p.traitement_id = tr.traitement_id
                                                   JOIN Contrat co ON tr.contrat_id = co.contrat_id
                                          WHERE co.client_id = %s
                                          ORDER BY annee DESC, mois DESC;
                                          """
                    await cursor.execute(query_client_months, (client_id_for_invoice,))
                    client_available_months = await cursor.fetchall()
        except Exception as e:
            print(f"Erreur lors de la récupération des mois disponibles pour le client : {e}")
            client_available_months = []

        annéeChoisi = None
        moisChoisi = None

        if client_available_months:
            print(f"\nMois de factures disponibles pour {client_full_name_for_invoice} :")
            for i, entry in enumerate(client_available_months):
                month_name = datetime.date(entry['annee'], entry['mois'], 1).strftime('%B').capitalize()
                print(f"  {i + 1}. {month_name} {entry['annee']}")
            print("  0. Entrer un autre mois/année manuellement (pour un mois non listé)")

            while True:
                try:
                    choice = int(
                        input("Choisissez un numéro de mois dans la liste ou '0' pour entrer manuellement : ").strip())
                    if 0 < choice <= len(client_available_months):
                        annéeChoisi = client_available_months[choice - 1]['annee']
                        moisChoisi = client_available_months[choice - 1]['mois']
                        break
                    elif choice == 0:
                        while True:
                            try:
                                year_input = input("Veuillez entrer l'année de la facture (ex: 2023) : ").strip()
                                month_input = input(
                                    "Veuillez entrer le numéro du mois (1-12) de la facture (ex: 6 pour Juin) : ").strip()

                                annéeChoisi = int(year_input)
                                moisChoisi = int(month_input)

                                if not (1 <= moisChoisi <= 12):
                                    print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                                    continue
                                if not (2000 <= annéeChoisi <= datetime.datetime.now().year + 5):
                                    print(
                                        f"Année invalide. Veuillez entrer une année entre 2000 et {datetime.datetime.now().year + 5}.")
                                    continue
                                break
                            except ValueError:
                                print("Entrée invalide. Veuillez entrer un nombre pour l'année et le mois.")
                        break
                    else:
                        print("Choix invalide. Veuillez réessayer.")
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un numéro.")
        else:
            print(
                f"\nAucune facture trouvée pour {client_full_name_for_invoice}. Veuillez entrer le mois et l'année manuellement.")
            while True:
                try:
                    year_input = input("Veuillez entrer l'année de la facture (ex: 2023) : ").strip()
                    month_input = input(
                        "Veuillez entrer le numéro du mois (1-12) de la facture (ex: 6 pour Juin) : ").strip()

                    annéeChoisi = int(year_input)
                    moisChoisi = int(month_input)

                    if not (1 <= moisChoisi <= 12):
                        print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                        continue
                    if not (2000 <= annéeChoisi <= datetime.datetime.now().year + 5):
                        print(
                            f"Année invalide. Veuillez entrer une année entre 2000 et {datetime.datetime.now().year + 5}.")
                        continue
                    break
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un nombre pour l'année et le mois.")

        if annéeChoisi is None or moisChoisi is None:
            print("Sélection de l'année ou du mois annulée. Fin du rapport.")
            return

        print(
            f"\nPréparation de la facture pour '{client_full_name_for_invoice}' pour {datetime.date(annéeChoisi, moisChoisi, 1).strftime('%B').capitalize()} {annéeChoisi}...")

        factures_data = await obtenirDataFactureClient(pool, client_id_for_invoice, annéeChoisi, moisChoisi)
        genererFactureExcel(factures_data, client_full_name_for_invoice, annéeChoisi, moisChoisi)

    except Exception as e:
        print(f"Une erreur inattendue est survenue dans le script principal : {e}")
    finally:
        if pool:
            pool.close()


if __name__ == "__main__":
    asyncio.run(generationFactureClient())