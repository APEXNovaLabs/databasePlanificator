import pandas as pd
import datetime
import asyncio
import aiomysql
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os # Import for file path handling

# Assurez-vous que ce chemin est correct pour votre configuration
# from Contrat.fonctionnalités.connexionDB import DBConnection
# Placeholder for DBConnection if not available, assuming it provides an aiomysql connection pool
async def DBConnection(host, user, password, db_name):
    try:
        # Create a connection pool using the provided credentials
        pool = await aiomysql.create_pool(
            host=host,
            user=user,
            password=password,
            db=db_name,
            autocommit=True,
            minsize=1,
            maxsize=5
        )
        print("Connexion à la base de données établie avec succès.")
        return pool
    except Exception as e:
        print(f"Erreur lors de la connexion à la base de données : {e}")
        return None


# --- Fonction de récupération des données de facture pour un client spécifique (pour rapport mensuel) ---
async def obtenirDataFactureClient(pool, client_id: int, year: int, month: int):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT cl.nom                  AS client_nom,
                           COALESCE(cl.prenom, '') AS client_prenom,
                           cl.adresse              AS client_adresse,
                           cl.telephone            AS client_telephone,
                           cl.categorie            AS client_categorie,
                           cl.axe                  AS client_axe,
                           co.reference_contrat    AS `Référence Contrat`,
                           f.reference_facture     AS `Numéro Facture`,
                           f.date_traitement       AS `Date de traitement`,
                           tt.typeTraitement       AS `Traitement (Type)`,
                           pd.statut               AS `Etat traitement`,
                           f.etat                  AS `Etat paiement (Payée ou non)`,
                           f.mode                  AS `Mode de Paiement`,
                           f.date_paiement         AS `Date de Paiement`,       -- Mise à jour: date_paiement pour toutes les modes
                           f.numero_cheque         AS `Numéro du Chèque`,       -- Nouveau
                           f.etablissemnt_payeur   AS `Établissement Payeur`,   -- Nouveau
                           COALESCE(
                                   (SELECT hp.new_amount
                                    FROM Historique_prix hp
                                    WHERE hp.facture_id = f.facture_id
                                    ORDER BY hp.change_date DESC, hp.history_id DESC
                                    LIMIT 1),
                                   f.montant
                           )                       AS montant_facture
                    FROM Facture f
                             JOIN PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                             JOIN Planning p ON pd.planning_id = p.planning_id
                             JOIN Traitement tr ON p.traitement_id = tr.traitement_id
                             JOIN TypeTraitement tt ON tr.id_type_traitement = tt.id_type_traitement
                             JOIN Contrat co ON tr.contrat_id = co.contrat_id
                             JOIN Client cl ON co.client_id = cl.client_id
                    WHERE cl.client_id = %s
                      AND YEAR(f.date_traitement) = %s
                      AND MONTH(f.date_traitement) = %s
                    ORDER BY f.date_traitement;
                    """
            await cursor.execute(query, (client_id, year, month))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des données de facture : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction de récupération des données de facture complètes pour un client (pour rapport annuel/complet) ---
async def get_factures_data_for_client_comprehensive(pool, client_id: int, start_date: datetime.date = None,
                                                     end_date: datetime.date = None):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT cl.nom                  AS client_nom,
                           COALESCE(cl.prenom, '') AS client_prenom,
                           cl.adresse              AS client_adresse,
                           cl.telephone            AS client_telephone,
                           cl.categorie            AS client_categorie,
                           cl.axe                  AS client_axe,
                           co.contrat_id,
                           co.reference_contrat    AS `Référence Contrat`,
                           co.date_contrat,
                           co.date_debut           AS contrat_date_debut,
                           co.date_fin             AS contrat_date_fin,
                           co.statut_contrat,
                           co.duree                AS contrat_duree_type,
                           f.reference_facture     AS `Numéro Facture`,
                           tt.typeTraitement       AS `Type de Traitement`,
                           pd.date_planification   AS `Date de Planification`,
                           pd.statut               AS `Etat du Planning`,
                           p.redondance            AS `Redondance (Mois)`,
                           f.date_traitement       AS `Date de Facturation`,
                           f.etat                  AS `Etat de Paiement`,
                           f.mode                  AS `Mode de Paiement`,
                           f.date_paiement         AS `Date de Paiement`,       -- Mise à jour: date_paiement pour toutes les modes
                           f.numero_cheque         AS `Numéro du Chèque`,       -- Nouveau
                           f.etablissemnt_payeur   AS `Établissement Payeur`,   -- Nouveau
                           COALESCE(
                                   (SELECT hp.new_amount
                                    FROM Historique_prix hp
                                    WHERE hp.facture_id = f.facture_id
                                    ORDER BY hp.change_date DESC, hp.history_id DESC
                                    LIMIT 1),
                                   f.montant
                           )                       AS `Montant Facturé`
                    FROM Client cl
                             JOIN Contrat co ON cl.client_id = co.client_id
                             JOIN Traitement tr ON co.contrat_id = tr.contrat_id
                             JOIN TypeTraitement tt ON tr.id_type_traitement = tt.id_type_traitement
                             JOIN Planning p ON tr.traitement_id = p.traitement_id
                             INNER JOIN PlanningDetails pd ON p.planning_id = pd.planning_id
                             INNER JOIN Facture f ON pd.planning_detail_id = f.planning_detail_id
                    WHERE cl.client_id = %s
                    """
            params = [client_id]

            if start_date and end_date:
                query += " AND f.date_traitement BETWEEN %s AND %s"
                params.append(start_date)
                params.append(end_date)
            elif start_date:
                query += " AND f.date_traitement >= %s"
                params.append(start_date)
            elif end_date:
                query += " AND f.date_traitement <= %s"
                params.append(end_date)

            query += " ORDER BY `Date de Planification` ASC, `Date de Facturation` ASC;"

            await cursor.execute(query, tuple(params))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des données de facture complètes : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction utilitaire pour obtenir le client_id et nom complet de tous les clients ---
async def obtenirTousClient(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id, CONCAT(nom, ' ', COALESCE(prenom, '')) AS full_name
                    FROM Client
                    ORDER BY full_name;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération de la liste des clients : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction utilitaire pour obtenir le client_id à partir du nom/prénom ---
async def obtenirIDClientAvecNom(pool, client_name: str):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id
                    FROM Client
                    WHERE CONCAT(nom, ' ', COALESCE(prenom, '')) = %s
                    LIMIT 1;
                    """
            await cursor.execute(query, (client_name,))
            result = await cursor.fetchone()
            return result['client_id'] if result else None
    except Exception as e:
        print(f"Erreur lors de la recherche du client par nom : {e}")
        return None
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour récupérer la liste des clients, le nombre de factures et les mois des factures ---
async def obtenirInformationsClients(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT c.client_id,
                           CONCAT(c.nom, ' ', COALESCE(c.prenom, '')) AS nomComplet,
                           COUNT(f.facture_id)                        AS total_factures,
                           GROUP_CONCAT(DISTINCT
                                        CONCAT(YEAR(f.date_traitement), '-', LPAD(MONTH(f.date_traitement), 2, '0'))
                                        ORDER BY YEAR(f.date_traitement) DESC, MONTH(f.date_traitement) DESC SEPARATOR
                                        ', ')                         AS facture_months
                    FROM Client c
                             LEFT JOIN Contrat co ON c.client_id = co.client_id
                             LEFT JOIN Traitement tr ON co.contrat_id = tr.contrat_id
                             LEFT JOIN Planning p ON tr.traitement_id = p.traitement_id
                             LEFT JOIN PlanningDetails pd ON p.planning_id = pd.planning_id
                             LEFT JOIN Facture f ON pd.planning_detail_id = f.planning_detail_id
                    GROUP BY c.client_id, nomComplet
                    ORDER BY nomComplet;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération du nombre de factures par client et des mois : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour obtenir la date de la première et de la dernière facture pour un client ---
async def get_client_earliest_latest_invoice_dates(pool, client_id: int):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT MIN(f.date_traitement) AS min_date,
                           MAX(f.date_traitement) AS max_date
                    FROM Facture f
                             JOIN PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                             JOIN Planning p ON pd.planning_id = p.planning_id
                             JOIN Traitement tr ON p.traitement_id = tr.traitement_id
                             JOIN Contrat co ON tr.contrat_id = co.contrat_id
                    WHERE co.client_id = %s;
                    """
            await cursor.execute(query, (client_id,))
            result = await cursor.fetchone()
            return result['min_date'], result['max_date']
    except Exception as e:
        print(f"Erreur lors de la récupération des dates de facture min/max : {e}")
        return None, None
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour générer le fichier Excel de la facture client (mensuel) ---
def genererFactureExcel(data: list[dict], client_full_name: str, year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()

    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"{safe_client_name}-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facture {client_full_name} {month_name_fr}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    ligneActuelle = 1

    # Informations du client (en-tête)
    if data:
        infoClient = data[0]
        affichageNomClient = f"{infoClient['client_nom']} {infoClient['client_prenom']}"
        if infoClient['client_categorie'] != 'Particulier':
            affichageNomClient = f"{infoClient['client_nom']} (Responsable: {infoClient['client_prenom'] if infoClient['client_prenom'] else 'N/A'})"

        ws.cell(row=ligneActuelle, column=1, value="Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=affichageNomClient)
        ligneActuelle += 1

        # Ajout du numéro de contrat
        ws.cell(row=ligneActuelle, column=1, value="N° Contrat :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient.get('Référence Contrat', 'N/A'))
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Adresse :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_adresse'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_telephone'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_categorie'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Axe Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_axe'])
        ligneActuelle += 1

    ligneActuelle += 1

    # Tableau des traitements
    table_headers = [
        'Numéro Facture', 'Date de Planification', 'Date de traitement', 'Traitement concerné',
        'Etat du Planning', 'Mode de Paiement', 'Détails Paiement', 'Etat de Paiement', 'Montant'
    ]
    num_table_cols = len(table_headers)

    # Ligne "Facture du mois de:"
    ws.cell(row=ligneActuelle, column=1, value=f"Facture du mois de : {month_name_fr} {year}").font = header_font
    ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=num_table_cols)
    ws.cell(row=ligneActuelle, column=1).alignment = Alignment(horizontal='center')
    ligneActuelle += 2

    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=ligneActuelle, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    ligneActuelle += 1

    if not data:
        ws.cell(row=ligneActuelle, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour ce mois.").border = thin_border
        ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=len(table_headers))
        ligneActuelle += 1
    else:
        # Convertir en DataFrame pour un traitement plus facile
        df_invoice_data = pd.DataFrame(data)

        for r_idx, row_dict in enumerate(df_invoice_data.to_dict('records'), start=ligneActuelle):
            # Gérer le numéro de facture: afficher "Aucun" si vide ou None
            invoice_number = row_dict.get('Numéro Facture')
            display_invoice_number = invoice_number if invoice_number else "Aucun"

            # Préparer les données de la ligne selon les en-têtes définis
            row_data = [
                display_invoice_number, # Utilisation de la valeur traitée
                row_dict.get('Date de Planification', 'N/A'),
                row_dict.get('Date de traitement', 'N/A'),
                row_dict.get('Traitement (Type)', 'N/A'),
                row_dict.get('Etat traitement', 'N/A'),
                row_dict.get('Mode de Paiement', 'N/A'),
                '',  # Placeholder for Détails Paiement
                row_dict.get('Etat paiement (Payée ou non)', 'N/A'),
                row_dict.get('montant_facture', 'N/A')
            ]

            # Gérer les détails de paiement
            mode_paiement = row_dict.get('Mode de Paiement')
            details_paiement = "N/A"
            date_paiement_obj = row_dict.get('Date de Paiement')
            date_paiement_str = date_paiement_obj.strftime('%Y-%m-%d') if date_paiement_obj else 'N/A'

            if mode_paiement == 'Chèque':
                numero_cheque_str = row_dict.get('Numéro du Chèque', 'N/A')
                etablissement_payeur_str = row_dict.get('Établissement Payeur', 'N/A')
                details_paiement = f"Chèque: {numero_cheque_str} ({date_paiement_str}, {etablissement_payeur_str})"
            elif mode_paiement == 'Virement':
                details_paiement = f"Virement: ({date_paiement_str})"
            elif mode_paiement == 'Mobile Money':
                details_paiement = "Mobile Money"
            elif mode_paiement == 'Espèce':
                details_paiement = "Paiement en espèces"
            row_data[6] = details_paiement # Mettre à jour la colonne 'Détails Paiement'

            payment_status = row_dict.get('Etat paiement (Payée ou non)')
            fill_to_apply = None
            if payment_status == 'Payé':
                fill_to_apply = green_fill
            elif payment_status == 'Non payé':
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            ligneActuelle += 1

    ligneActuelle += 1

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)

        total_by_type_paid = df_calc[df_calc['Etat paiement (Payée ou non)'] == 'Payé'].groupby('Traitement (Type)')[
            'montant_facture'].sum()

        if not total_by_type_paid.empty:
            ws.cell(row=ligneActuelle, column=1, value="Facture total pour :").font = bold_font
            ligneActuelle += 1
            for service_type, total_amount in total_by_type_paid.items():
                ws.cell(row=ligneActuelle, column=2, value=f"{service_type} (Payé)").font = bold_font
                ws.cell(row=ligneActuelle, column=3, value=total_amount).font = bold_font
                ligneActuelle += 1
        else:
            ws.cell(row=ligneActuelle, column=1,
                    value="Aucun montant payé pour les types de traitement ce mois.").font = bold_font
            ligneActuelle += 1

        ligneActuelle += 1

        # Total de paiement par mode de paiement
        ws.cell(row=ligneActuelle, column=1, value="Total de paiement par mode de paiement :").font = bold_font
        ligneActuelle += 1
        payment_mode_counts = df_calc.groupby('Mode de Paiement').size().reset_index(name='Nombre de Paiements')
        for _, row in payment_mode_counts.iterrows():
            ws.cell(row=ligneActuelle, column=2, value=f"{row['Mode de Paiement']} :").font = bold_font
            ws.cell(row=ligneActuelle, column=3, value=row['Nombre de Paiements']).font = bold_font
            ligneActuelle += 1
        ligneActuelle += 1

        grand_total = df_calc['montant_facture'].sum()
        ws.cell(row=ligneActuelle, column=1, value="Montant total des traitements effectués ce mois :").font = bold_font
        ws.cell(row=ligneActuelle, column=3, value=grand_total).font = bold_font
        ligneActuelle += 1

    max_col_for_width = len(table_headers)

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                # Gérer les dates pour le calcul de la largeur
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        # Save to a file in the current directory
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")


# --- Fonction pour générer le fichier Excel de la facture client complète (annuel/complet) ---
def generate_comprehensive_facture_excel(data: list[dict], client_full_name: str, report_period: str):
    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"Rapport_Factures_{safe_client_name}_{report_period.replace(' ', '_')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Factures {client_full_name} {report_period}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    current_row = 1

    # Informations du client (en-tête)
    if data:
        client_info = data[0]
        client_display_name = f"{client_info['client_nom']} {client_info['client_prenom']}"
        if client_info['client_categorie'] != 'Particulier':
            client_display_name = f"{client_info['client_nom']} (Responsable: {client_info['client_prenom'] if client_info['client_prenom'] else 'N/A'})"

        ws.cell(row=current_row, column=1, value="Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_display_name)
        current_row += 1

        # Ajout du numéro de contrat
        ws.cell(row=current_row, column=1, value="N° Contrat :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info.get('Référence Contrat', 'N/A'))
        current_row += 1

        ws.cell(row=current_row, column=1, value="Adresse :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_adresse'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_telephone'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_categorie'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Axe Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_axe'])
        current_row += 1

    current_row += 1

    # Ligne de titre du rapport
    ws.cell(row=current_row, column=1,
            value=f"Rapport de Facturation pour la période : {report_period}").font = header_font
    table_headers = [
        'Numéro Facture', 'Date de Planification', 'Date de Facturation', 'Type de Traitement',
        'Etat du Planning', 'Mode de Paiement', 'Détails Paiement', 'Etat de Paiement', 'Montant Facturé'
    ]
    max_cols_for_merge = len(table_headers)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols_for_merge)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    current_row += 2

    # Tableau des détails de la facture
    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    current_row += 1

    if not data:
        ws.cell(row=current_row, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour la période sélectionnée.").border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(table_headers))
        current_row += 1
    else:
        df_invoice_data = pd.DataFrame(data)

        for r_idx, row_dict in enumerate(df_invoice_data.to_dict('records'), start=current_row):
            # Gérer le numéro de facture: afficher "Aucun" si vide ou None
            invoice_number = row_dict.get('Numéro Facture')
            display_invoice_number = invoice_number if invoice_number else "Aucun"

            row_data = [
                display_invoice_number, # Utilisation de la valeur traitée
                row_dict.get('Date de Planification', 'N/A'),
                row_dict.get('Date de Facturation', 'N/A'),
                row_dict.get('Type de Traitement', 'N/A'),
                row_dict.get('Etat du Planning', 'N/A'),
                row_dict.get('Mode de Paiement', 'N/A'),
                '',  # Placeholder for Détails Paiement
                row_dict.get('Etat de Paiement', 'N/A'),
                row_dict.get('Montant Facturé', 'N/A')
            ]

            # Gérer les détails de paiement
            mode_paiement = row_dict.get('Mode de Paiement')
            details_paiement = "N/A"
            date_paiement_obj = row_dict.get('Date de Paiement')
            date_paiement_str = date_paiement_obj.strftime('%Y-%m-%d') if date_paiement_obj else 'N/A'

            if mode_paiement == 'Chèque':
                numero_cheque_str = row_dict.get('Numéro du Chèque', 'N/A')
                etablissement_payeur_str = row_dict.get('Établissement Payeur', 'N/A')
                details_paiement = f"Chèque: {numero_cheque_str} ({date_paiement_str}, {etablissement_payeur_str})"
            elif mode_paiement == 'Virement':
                details_paiement = f"Virement: ({date_paiement_str})"
            elif mode_paiement == 'Mobile Money':
                details_paiement = "Mobile Money"
            elif mode_paiement == 'Espèce':
                details_paiement = "Paiement en espèces"
            row_data[6] = details_paiement # Mettre à jour la colonne 'Détails Paiement'


            payment_status = row_dict.get('Etat de Paiement')
            fill_to_apply = None
            if payment_status == 'Payé':
                fill_to_apply = green_fill
            elif payment_status == 'Non payé':
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            current_row += 1

    current_row += 1

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)

        grand_total = df_calc['Montant Facturé'].sum()
        ws.cell(row=current_row, column=1, value="Montant Total Facturé sur la période :").font = bold_font
        ws.cell(row=current_row, column=len(table_headers), value=grand_total).font = bold_font
        current_row += 1

        total_paid = df_calc[df_calc['Etat de Paiement'] == 'Payé']['Montant Facturé'].sum()
        ws.cell(row=current_row, column=1, value="Montant Total Payé sur la période :").font = bold_font
        ws.cell(row=current_row, column=len(table_headers), value=total_paid).font = bold_font
        ws.cell(row=current_row, column=len(table_headers)).fill = green_fill
        current_row += 1

        total_unpaid = df_calc[df_calc['Etat de Paiement'] == 'Non payé']['Montant Facturé'].sum()
        ws.cell(row=current_row, column=1, value="Montant Total Impayé sur la période :").font = bold_font
        ws.cell(row=current_row, column=len(table_headers), value=total_unpaid).font = bold_font
        ws.cell(row=current_row, column=len(table_headers)).fill = red_fill
        current_row += 1

        current_row += 1

        # Total de paiement par mode de paiement
        ws.cell(row=current_row, column=1, value="Total de paiement par mode de paiement :").font = bold_font
        current_row += 1
        payment_mode_counts = df_calc.groupby('Mode de Paiement').size().reset_index(name='Nombre de Paiements')
        for _, row in payment_mode_counts.iterrows():
            ws.cell(row=current_row, column=2, value=f"{row['Mode de Paiement']} :").font = bold_font
            ws.cell(row=current_row, column=3, value=row['Nombre de Paiements']).font = bold_font
            current_row += 1
        current_row += 1

        ws.cell(row=current_row, column=1, value="Synthèse par Type de Traitement :").font = bold_font
        current_row += 1

        if 'Type de Traitement' in df_calc.columns:
            total_by_type = df_calc.groupby('Type de Traitement')['Montant Facturé'].sum().reset_index()
            for _, row in total_by_type.iterrows():
                ws.cell(row=current_row, column=2, value=row['Type de Traitement']).font = bold_font
                ws.cell(row=current_row, column=3, value=row['Montant Facturé']).font = bold_font
                current_row += 1
        else:
            ws.cell(row=current_row, column=1,
                    value="Impossible de synthétiser par type de traitement (colonne manquante).")
            current_row += 1

    for i in range(1, len(table_headers) + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                # Gérer les dates pour le calcul de la largeur
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        # Save to a file in the current directory
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")


# --- Fonction principale pour exécuter la génération de rapport de facture ---
async def main_invoice_report_menu():
    pool = None
    try:
        print("\n--- Configuration de la connexion à la base de données ---")
        db_host = input("Entrez l'hôte de la base de données (ex: localhost): ")
        db_user = input("Entrez le nom d'utilisateur de la base de données (ex: root): ")
        db_password = input("Entrez le mot de passe de la base de données: ")
        db_name = input("Entrez le nom de la base de données (ex: Planificator): ")

        pool = await DBConnection(db_host, db_user, db_password, db_name)
        if pool is None:
            print("Échec de la connexion à la base de données. Annulation de l'opération.")
            return

        continue_generating_reports = True
        while continue_generating_reports:
            print("\n--- Aperçu des clients et de leurs factures ---")
            client_infos = await obtenirInformationsClients(pool)
            if not client_infos:
                print("Aucun client trouvé dans la base de données.")
                break

            print("\nListe des clients et leurs informations de facture:")
            for i, client in enumerate(client_infos):
                print(f"{i + 1}. {client['nomComplet']} (Factures: {client['total_factures']}, Mois: {client['facture_months'] if client['facture_months'] else 'Aucun'})")

            print("\n--- Menu de Génération de Rapports ---")
            print("1. Générer un rapport mensuel pour un client")
            print("2. Générer un rapport complet/annuel pour un client")
            print("3. Quitter")

            choice = input("Entrez votre choix (1-3): ")

            if choice == '1':
                client_index = input("Entrez le numéro du client pour le rapport mensuel: ")
                try:
                    client_index = int(client_index) - 1
                    if 0 <= client_index < len(client_infos):
                        selected_client = client_infos[client_index]
                        client_id = selected_client['client_id']
                        client_full_name = selected_client['nomComplet']

                        year = int(input("Entrez l'année (ex: 2023): "))
                        month = int(input("Entrez le mois (1-12): "))

                        print(f"Récupération des données de facture pour {client_full_name} ({month}/{year})...")
                        invoice_data = await obtenirDataFactureClient(pool, client_id, year, month)
                        if invoice_data:
                            print(f"Génération du rapport mensuel pour {client_full_name}...")
                            genererFactureExcel(invoice_data, client_full_name, year, month)
                        else:
                            print(f"Aucune donnée de facture trouvée pour {client_full_name} en {month}/{year}.")
                    else:
                        print("Numéro de client invalide.")
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un nombre.")
                except Exception as e:
                    print(f"Une erreur inattendue est survenue : {e}")

            elif choice == '2':
                client_index = input("Entrez le numéro du client pour le rapport complet/annuel: ")
                try:
                    client_index = int(client_index) - 1
                    if 0 <= client_index < len(client_infos):
                        selected_client = client_infos[client_index]
                        client_id = selected_client['client_id']
                        client_full_name = selected_client['nomComplet']

                        min_date, max_date = await get_client_earliest_latest_invoice_dates(pool, client_id)
                        print(f"Dates de facture disponibles pour {client_full_name}:")
                        print(f"  Première facture: {min_date if min_date else 'N/A'}")
                        print(f"  Dernière facture: {max_date if max_date else 'N/A'}")

                        report_type = input("Générer pour (A)nnée spécifique, (P)ériode spécifique, ou (T)outes les données? (A/P/T): ").upper()
                        start_date = None
                        end_date = None
                        report_period_str = "Toutes les données"

                        if report_type == 'A':
                            year_input = int(input("Entrez l'année pour le rapport annuel (ex: 2023): "))
                            start_date = datetime.date(year_input, 1, 1)
                            end_date = datetime.date(year_input, 12, 31)
                            report_period_str = f"Année {year_input}"
                        elif report_type == 'P':
                            start_date_str = input("Entrez la date de début (AAAA-MM-JJ): ")
                            end_date_str = input("Entrez la date de fin (AAAA-MM-JJ): ")
                            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
                            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
                            report_period_str = f"Du {start_date_str} au {end_date_str}"
                        elif report_type == 'T':
                            pass # No date filters, get all data
                        else:
                            print("Choix de période invalide.")
                            continue

                        print(f"Récupération des données complètes pour {client_full_name}...")
                        comprehensive_data = await get_factures_data_for_client_comprehensive(pool, client_id, start_date, end_date)
                        if comprehensive_data:
                            print(f"Génération du rapport complet pour {client_full_name}...")
                            generate_comprehensive_facture_excel(comprehensive_data, client_full_name, report_period_str)
                        else:
                            print(f"Aucune donnée de facture complète trouvée pour {client_full_name} pour la période sélectionnée.")
                    else:
                        print("Numéro de client invalide.")
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un nombre ou un format de date valide.")
                except Exception as e:
                    print(f"Une erreur inattendue est survenue : {e}")

            elif choice == '3':
                print("Quitter le programme.")
                continue_generating_reports = False
            else:
                print("Choix invalide. Veuillez réessayer.")

    except Exception as e:
        print(f"Une erreur critique est survenue dans le menu principal : {e}")
    finally:
        if pool:
            pool.close()
            print("Connexion à la base de données fermée.")

# Pour exécuter le menu principal
if __name__ == "__main__":
    # Ensure the directory for generated files exists
    if not os.path.exists("./generated_invoices"):
        os.makedirs("./generated_invoices")
    os.chdir("./generated_invoices") # Change to this directory to save files there

    asyncio.run(main_invoice_report_menu())
