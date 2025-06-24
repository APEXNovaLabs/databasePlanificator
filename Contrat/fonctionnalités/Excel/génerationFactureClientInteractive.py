import pandas as pd
import datetime
import asyncio
import aiomysql
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from Contrat.fonctionnalités.connexionDB import DBConnection

# --- Fonction de récupération des données de facture pour un client spécifique ---
async def get_factures_data_for_client(pool, client_id: int, year: int, month: int):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT cl.nom            AS client_nom,
                           cl.prenom         AS client_prenom,
                           cl.adresse        AS client_adresse,
                           cl.telephone      AS client_telephone,
                           cl.categorie      AS client_categorie,
                           f.date_traitement AS `Date de traitement`,
                           tt.typeTraitement AS `Traitement (Type)`,
                           pd.statut         AS `Etat traitement`,
                           f.etat            AS `Etat paiement (Payée ou non)`,
                           f.montant         AS montant_facture
                    FROM Facture f
                             JOIN
                         PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                             JOIN
                         Planning p ON pd.planning_id = p.planning_id
                             JOIN
                         Traitement tr ON p.traitement_id = tr.traitement_id
                             JOIN
                         TypeTraitement tt ON tr.id_type_traitement = tt.id_type_traitement
                             JOIN
                         Contrat co ON tr.contrat_id = co.contrat_id
                             JOIN
                         Client cl ON co.client_id = cl.client_id
                    WHERE cl.client_id = %s
                      AND YEAR(f.date_traitement) = %s
                      AND MONTH(f.date_traitement) = %s
                    ORDER BY f.date_traitement;
                    """
            await cursor.execute(query, (client_id, year, month))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des données de facture : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction utilitaire pour obtenir le client_id et nom complet de tous les clients ---
async def get_all_clients(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id, CONCAT(nom, ' ', prenom) AS full_name
                    FROM Client
                    ORDER BY full_name;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération de la liste des clients : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction utilitaire pour obtenir le client_id à partir du nom/prénom ---
async def get_client_id_by_name(pool, client_name: str):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id
                    FROM Client
                    WHERE CONCAT(nom, ' ', prenom) = %s
                    LIMIT 1;
                    """
            await cursor.execute(query, (client_name,))
            result = await cursor.fetchone()
            return result['client_id'] if result else None
    except Exception as e:
        print(f"Erreur lors de la recherche du client par nom : {e}")
        return None
    finally:
        if conn:
            pool.release(conn)

# --- Fonction pour récupérer la liste des clients et le nombre de factures par client ---
async def get_client_invoice_counts(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT
                        c.client_id,
                        CONCAT(c.nom, ' ', c.prenom) AS full_name,
                        COUNT(f.facture_id) AS total_factures
                    FROM Client c
                    LEFT JOIN Contrat co ON c.client_id = co.client_id
                    LEFT JOIN Traitement tr ON co.contrat_id = tr.contrat_id
                    LEFT JOIN Planning p ON tr.traitement_id = p.traitement_id
                    LEFT JOIN PlanningDetails pd ON p.planning_id = pd.planning_id
                    LEFT JOIN Facture f ON pd.planning_detail_id = f.planning_detail_id
                    GROUP BY c.client_id, full_name
                    ORDER BY full_name;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération du nombre de factures par client : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction pour générer le fichier Excel de la facture client ---
def generate_facture_excel(data: list[dict], client_full_name: str, year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()

    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"{safe_client_name}-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facture {client_full_name} {month_name_fr}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    current_row = 1

    # Informations du client (en-tête)
    if data:
        client_info = data[0]
        client_display_name = f"{client_info['client_nom']} {client_info['client_prenom']}"
        if client_info['client_categorie'] != 'Particulier':
            client_display_name = f"{client_info['client_nom']} (Responsable: {client_info['client_prenom']})"

        ws.cell(row=current_row, column=1, value="Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_display_name)
        current_row += 1

        ws.cell(row=current_row, column=1, value="Adresse :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_adresse'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_telephone'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_categorie'])
        current_row += 1

    current_row += 1  # Ligne vide

    # Ligne "Facture du mois de:"
    num_table_cols = 4  # Fixed number of columns for the main invoice table
    ws.cell(row=current_row, column=1, value=f"Facture du mois de : {month_name_fr} {year}").font = header_font
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_table_cols)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    current_row += 2  # Deux lignes vides après pour la séparation avec le tableau

    # Tableau des traitements
    table_headers = ['Date de traitement', 'Traitement (Type)', 'Etat traitement', 'Etat paiement (Payée ou non)']

    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    current_row += 1

    if not data:
        ws.cell(row=current_row, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour ce mois.").border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(table_headers))
        current_row += 1
    else:
        df_invoice_data = pd.DataFrame(data)
        # Sélectionner et réordonner les colonnes pour l'affichage du tableau
        df_display = df_invoice_data[
            ['Date de traitement', 'Traitement (Type)', 'Etat traitement', 'Etat paiement (Payée ou non)']]

        # Écrire les données du tableau
        for r_idx, row_data in enumerate(df_display.values.tolist(), start=current_row):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
            current_row += 1

    current_row += 1  # Ligne vide avant les totaux

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)  # Utilise le DataFrame complet avec montant_facture

        # Total par type de traitement (seulement payé)
        total_by_type_paid = df_calc[df_calc['Etat paiement (Payée ou non)'] == 'Payé'].groupby('Traitement (Type)')[
            'montant_facture'].sum()

        if not total_by_type_paid.empty:
            ws.cell(row=current_row, column=1, value="Facture total pour :").font = bold_font
            current_row += 1
            for service_type, total_amount in total_by_type_paid.items():
                ws.cell(row=current_row, column=2, value=f"{service_type} (Payé)").font = bold_font
                ws.cell(row=current_row, column=3, value=total_amount).font = bold_font
                current_row += 1
        else:
            ws.cell(row=current_row, column=1,
                    value="Aucun montant payé pour les types de traitement ce mois.").font = bold_font
            current_row += 1

        current_row += 1  # Ligne vide

        # Montant total pour tous les traitements effectués (même non payés)
        grand_total = df_calc['montant_facture'].sum()
        ws.cell(row=current_row, column=1, value="Montant total des traitements effectués ce mois :").font = bold_font
        ws.cell(row=current_row, column=3, value=grand_total).font = bold_font
        current_row += 1

    # Ajuster la largeur des colonnes
    max_col_for_width = max(num_table_cols, 3)

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")

# --- Fonction principale pour exécuter la génération de facture ---
async def main_client_invoice():
    pool = None
    try:
        pool = await DBConnection()
        if not pool:
            print("Échec de la connexion à la base de données. Annulation de l'opération.")
            return

        # --- Affichage du tableau des clients et de leurs nombres de factures ---
        print("\n--- Aperçu des clients et de leurs factures ---")
        client_counts = await get_client_invoice_counts(pool)

        if client_counts:
            # Créer un DataFrame pour un affichage tabulaire clair
            df_counts = pd.DataFrame(client_counts)
            df_counts.rename(columns={'full_name': 'Nom du Client', 'total_factures': 'Nombre de Factures'}, inplace=True)
            df_counts.index = df_counts.index + 1 # Pour commencer l'index à 1
            print(df_counts[['Nom du Client', 'Nombre de Factures']].to_string())
        else:
            print("Aucun client trouvé ou aucune facture enregistrée.")

        # --- Suite du script pour la sélection du client et la génération de facture ---
        current_year = datetime.datetime.now().year
        current_month = datetime.datetime.now().month

        print("\n--- Génération de facture client ---")
        clients = await get_all_clients(pool)
        if not clients:
            print("Aucun client trouvé dans la base de données. Impossible de générer une facture.")
            return

        client_map = {}
        for i, client in enumerate(clients):
            client_map[str(i + 1)] = client # Store client_id for selection by number

        # Display the numbered list of clients for selection (already done by the df_counts, but for consistency if df_counts is skipped)
        # Re-list for selection if the user wants to pick by number directly after seeing the count table
        print("\nSélectionnez un client pour générer sa facture:")
        for i, client in enumerate(clients):
             print(f"{i + 1}. {client['full_name']}")


        client_id_for_invoice = None
        client_full_name_for_invoice = None

        while client_id_for_invoice is None:
            choice = input(
                "\nVeuillez entrer le numéro du client dans la liste, ou son nom complet (Nom Prénom) : ").strip()

            if choice.isdigit():
                if choice in client_map:
                    selected_client = client_map[choice]
                    client_id_for_invoice = selected_client['client_id']
                    client_full_name_for_invoice = selected_client['full_name']
                    print(f"Client sélectionné : {client_full_name_for_invoice}")
                else:
                    print("Numéro invalide. Veuillez réessayer.")
            else:
                client_full_name_for_invoice = choice
                client_id_for_invoice = await get_client_id_by_name(pool,
                                                                    client_full_name_for_invoice)
                if client_id_for_invoice is None:
                    print(f"Client '{client_full_name_for_invoice}' non trouvé. Veuillez vérifier le nom et réessayer.")
                else:
                    print(f"Client trouvé : {client_full_name_for_invoice}")

        print(
            f"\nRécupération des données de facture pour '{client_full_name_for_invoice}' pour {datetime.date(current_year, current_month, 1).strftime('%B').capitalize()} {current_year}...")

        factures_data = await get_factures_data_for_client(pool, client_id_for_invoice, current_year, current_month)
        generate_facture_excel(factures_data, client_full_name_for_invoice, current_year, current_month)

    except Exception as e:
        print(f"Une erreur inattendue est survenue dans le script principal : {e}")
    finally:
        if pool:
            pool.close()


if __name__ == "__main__":
    asyncio.run(main_client_invoice())