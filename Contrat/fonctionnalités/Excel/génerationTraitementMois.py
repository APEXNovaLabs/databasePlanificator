import pandas as pd
import datetime
import asyncio
import aiomysql
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

async def DBConnection(host, user, password, db_name):
    """
    Établit une connexion à la base de données MySQL et crée un pool de connexions.
    """
    try:
        pool = await aiomysql.create_pool(
            host=host,
            user=user,
            password=password,
            db=db_name,
            autocommit=True,
            minsize=1,
            maxsize=5
        )
        print("Connexion à la base de données établie avec succès.")
        return pool
    except Exception as e:
        print(f"Erreur lors de la connexion à la base de données : {e}")
        return None


# --- Fonction de récupération des données de facture pour un client spécifique (pour rapport mensuel - ANCIENNE FONCTION, NON UTILISÉE DANS LE NOUVEAU MENU) ---
# Cette fonction n'est plus directement appelée par le nouveau menu principal,
# mais elle est conservée pour référence si nécessaire.
async def obtenirDataFactureClient(pool, client_id: int, year: int, month: int):
    """
    Récupère les données de facture pour un client spécifique pour un mois et une année donnés.
    """
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT cl.nom                  AS client_nom,
                           COALESCE(cl.prenom, '') AS client_prenom,
                           cl.adresse              AS client_adresse,
                           cl.telephone            AS client_telephone,
                           cl.categorie            AS client_categorie,
                           cl.axe                  AS client_axe,
                           co.reference_contrat    AS `Référence Contrat`,
                           f.reference_facture     AS `Numéro Facture`,
                           f.date_traitement       AS `Date de traitement`,
                           tt.typeTraitement       AS `Traitement (Type)`,
                           pd.statut               AS `Etat traitement`,
                           f.etat                  AS `Etat paiement (Payée ou non)`,
                           f.mode                  AS `Mode de Paiement`,
                           f.date_paiement         AS `Date de Paiement`,
                           f.numero_cheque         AS `Numéro du Chèque`,
                           f.etablissemnt_payeur   AS `Établissement Payeur`,
                           COALESCE(
                                   (SELECT hp.new_amount
                                    FROM Historique_prix hp
                                    WHERE hp.facture_id = f.facture_id
                                    ORDER BY hp.change_date DESC, hp.history_id DESC
                                    LIMIT 1),
                                   f.montant
                           )                       AS montant_facture
                    FROM Facture f
                             JOIN PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                             JOIN Planning p ON pd.planning_id = p.planning_id
                             JOIN Traitement tr ON p.traitement_id = tr.traitement_id
                             JOIN TypeTraitement tt ON tr.id_type_traitement = tt.id_type_traitement
                             JOIN Contrat co ON tr.contrat_id = co.contrat_id
                             JOIN Client cl ON co.client_id = cl.client_id
                    WHERE cl.client_id = %s
                      AND YEAR(f.date_traitement) = %s
                      AND MONTH(f.date_traitement) = %s
                    ORDER BY f.date_traitement;
                    """
            await cursor.execute(query, (client_id, year, month))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des données de facture : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction de récupération des données de traitement complètes pour un client (pour rapport annuel/complet) ---
async def get_treatments_data_for_client_comprehensive(pool, client_id: int, start_date: datetime.date = None,
                                                     end_date: datetime.date = None):
    """
    Récupère toutes les données de traitement pour un client spécifique,
    éventuellement filtrées par une plage de dates de planification.
    Inclut les remarques et signalements liés aux planifications.
    """
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT cl.nom AS client_nom,
                           COALESCE(cl.prenom, '') AS client_prenom,
                           cl.adresse AS client_adresse,
                           cl.telephone AS client_telephone,
                           cl.categorie AS client_categorie,
                           cl.axe AS client_axe,
                           co.reference_contrat AS `Référence Contrat`,
                           co.date_contrat,
                           co.date_debut AS contrat_date_debut,
                           co.date_fin AS contrat_date_fin,
                           co.statut_contrat,
                           co.duree AS contrat_duree_type,
                           tt.typeTraitement AS `Type de Traitement`,
                           p.redondance AS `Redondance (Mois)`,
                           pd.date_planification AS `Date de Planification`,
                           pd.statut AS `Statut du Planning`,
                           r.contenu AS `Remarque`,
                           s.motif AS `Motif Signalement`,
                           s.type AS `Type Signalement`
                    FROM Client cl
                    JOIN Contrat co ON cl.client_id = co.client_id
                    JOIN Traitement tr ON co.contrat_id = tr.contrat_id
                    JOIN TypeTraitement tt ON tr.id_type_traitement = tt.id_type_traitement
                    JOIN Planning p ON tr.traitement_id = p.traitement_id
                    JOIN PlanningDetails pd ON p.planning_id = pd.planning_id
                    LEFT JOIN Remarque r ON pd.planning_detail_id = r.planning_detail_id
                    LEFT JOIN Signalement s ON pd.planning_detail_id = s.planning_detail_id
                    WHERE cl.client_id = %s
                    """
            params = [client_id]

            if start_date and end_date:
                query += " AND pd.date_planification BETWEEN %s AND %s"
                params.append(start_date)
                params.append(end_date)
            elif start_date:
                query += " AND pd.date_planification >= %s"
                params.append(start_date)
            elif end_date:
                query += " AND pd.date_planification <= %s"
                params.append(end_date)

            query += " ORDER BY `Date de Planification` ASC;"

            await cursor.execute(query, tuple(params))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des données de traitement complètes : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction utilitaire pour obtenir le client_id et nom complet de tous les clients ---
async def obtenirTousClient(pool):
    """
    Récupère l'ID client et le nom complet de tous les clients.
    """
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id, CONCAT(nom, ' ', COALESCE(prenom, '')) AS full_name
                    FROM Client
                    ORDER BY full_name;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération de la liste des clients : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction utilitaire pour obtenir le client_id à partir du nom/prénom ---
async def obtenirIDClientAvecNom(pool, client_name: str):
    """
    Récupère l'ID d'un client à partir de son nom complet.
    """
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT client_id
                    FROM Client
                    WHERE CONCAT(nom, ' ', COALESCE(prenom, '')) = %s
                    LIMIT 1;
                    """
            await cursor.execute(query, (client_name,))
            result = await cursor.fetchone()
            return result['client_id'] if result else None
    except Exception as e:
        print(f"Erreur lors de la recherche du client par nom : {e}")
        return None
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour récupérer la liste des clients, le nombre de factures et les mois des factures ---
async def obtenirInformationsClients(pool):
    """
    Récupère des informations agrégées sur les clients, y compris le nombre de factures
    et les mois de facturation.
    """
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT c.client_id,
                           CONCAT(c.nom, ' ', COALESCE(c.prenom, '')) AS nomComplet,
                           COUNT(f.facture_id)                        AS total_factures,
                           GROUP_CONCAT(DISTINCT
                                        CONCAT(YEAR(f.date_traitement), '-', LPAD(MONTH(f.date_traitement), 2, '0'))
                                        ORDER BY YEAR(f.date_traitement) DESC, MONTH(f.date_traitement) DESC SEPARATOR
                                        ', ')                         AS facture_months
                    FROM Client c
                             LEFT JOIN Contrat co ON c.client_id = co.client_id
                             LEFT JOIN Traitement tr ON co.contrat_id = tr.contrat_id
                             LEFT JOIN Planning p ON tr.traitement_id = p.traitement_id
                             LEFT JOIN PlanningDetails pd ON p.planning_id = pd.planning_id
                             LEFT JOIN Facture f ON pd.planning_detail_id = f.planning_detail_id
                    GROUP BY c.client_id, nomComplet
                    ORDER BY nomComplet;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération du nombre de factures par client et des mois : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour obtenir la date de la première et de la dernière facture pour un client ---
async def get_client_earliest_latest_invoice_dates(pool, client_id: int):
    """
    Récupère la date de la première et de la dernière facture pour un client donné.
    """
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT MIN(f.date_traitement) AS min_date,
                           MAX(f.date_traitement) AS max_date
                    FROM Facture f
                             JOIN PlanningDetails pd ON f.planning_detail_id = pd.planning_detail_id
                             JOIN Planning p ON pd.planning_id = p.planning_id
                             JOIN Traitement tr ON p.traitement_id = tr.traitement_id
                             JOIN Contrat co ON tr.contrat_id = co.contrat_id
                    WHERE co.client_id = %s;
                    """
            await cursor.execute(query, (client_id,))
            result = await cursor.fetchone()
            return result['min_date'], result['max_date']
    except Exception as e:
        print(f"Erreur lors de la récupération des dates de facture min/max : {e}")
        return None, None
    finally:
        if conn:
            pool.release(conn)


# --- Fonction pour générer le fichier Excel de la facture client (mensuel - ANCIENNE FONCTION, NON UTILISÉE DANS LE NOUVEAU MENU) ---
# Cette fonction n'est plus directement appelée par le nouveau menu principal,
# mais elle est conservée pour référence si nécessaire.
def genererFactureExcel(data: list[dict], client_full_name: str, year: int, month: int):
    """
    Génère un fichier Excel pour un rapport de facture mensuel d'un client.
    """
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()

    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"{safe_client_name}-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facture {client_full_name} {month_name_fr}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    ligneActuelle = 1

    # Informations du client (en-tête)
    if data:
        infoClient = data[0]
        affichageNomClient = f"{infoClient['client_nom']} {infoClient['client_prenom']}"
        if infoClient['client_categorie'] != 'Particulier':
            affichageNomClient = f"{infoClient['client_nom']} (Responsable: {infoClient['client_prenom'] if infoClient['client_prenom'] else 'N/A'})"

        ws.cell(row=ligneActuelle, column=1, value="Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=affichageNomClient)
        ligneActuelle += 1

        # Ajout du numéro de contrat
        ws.cell(row=ligneActuelle, column=1, value="N° Contrat :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient.get('Référence Contrat', 'N/A'))
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Adresse :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_adresse'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_telephone'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_categorie'])
        ligneActuelle += 1

        ws.cell(row=ligneActuelle, column=1, value="Axe Client :").font = bold_font
        ws.cell(row=ligneActuelle, column=2, value=infoClient['client_axe'])
        ligneActuelle += 1

    ligneActuelle += 1

    # Tableau des traitements
    table_headers = [
        'Numéro Facture', 'Date de Planification', 'Date de traitement', 'Traitement concerné',
        'Etat du Planning', 'Mode de Paiement', 'Détails Paiement', 'Etat de Paiement', 'Montant'
    ]
    num_table_cols = len(table_headers)

    # Ligne "Facture du mois de:"
    ws.cell(row=ligneActuelle, column=1, value=f"Facture du mois de : {month_name_fr} {year}").font = header_font
    ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=num_table_cols)
    ws.cell(row=ligneActuelle, column=1).alignment = Alignment(horizontal='center')
    ligneActuelle += 2

    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=ligneActuelle, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    ligneActuelle += 1

    if not data:
        ws.cell(row=ligneActuelle, column=1,
                value=f"Aucune facture trouvée pour le client '{client_full_name}' pour ce mois.").border = thin_border
        ws.merge_cells(start_row=ligneActuelle, start_column=1, end_row=ligneActuelle, end_column=len(table_headers))
        ligneActuelle += 1
    else:
        # Convertir en DataFrame pour un traitement plus facile
        df_invoice_data = pd.DataFrame(data)

        for r_idx, row_dict in enumerate(df_invoice_data.to_dict('records'), start=ligneActuelle):
            # Gérer le numéro de facture: afficher "Aucun" si vide ou None
            invoice_number = row_dict.get('Numéro Facture')
            display_invoice_number = invoice_number if invoice_number else "Aucun"

            # Préparer les données de la ligne selon les en-têtes définis
            row_data = [
                display_invoice_number, # Utilisation de la valeur traitée
                row_dict.get('Date de Planification', 'N/A'),
                row_dict.get('Date de traitement', 'N/A'),
                row_dict.get('Traitement (Type)', 'N/A'),
                row_dict.get('Etat traitement', 'N/A'),
                row_dict.get('Mode de Paiement', 'N/A'),
                '',  # Placeholder for Détails Paiement
                row_dict.get('Etat paiement (Payée ou non)', 'N/A'),
                row_dict.get('montant_facture', 'N/A')
            ]

            # Gérer les détails de paiement
            mode_paiement = row_dict.get('Mode de Paiement')
            details_paiement = "N/A"
            date_paiement_obj = row_dict.get('Date de Paiement')
            date_paiement_str = date_paiement_obj.strftime('%Y-%m-%d') if date_paiement_obj else 'N/A'

            if mode_paiement == 'Chèque':
                numero_cheque_str = row_dict.get('Numéro du Chèque', 'N/A')
                etablissement_payeur_str = row_dict.get('Établissement Payeur', 'N/A')
                details_paiement = f"Chèque: {numero_cheque_str} ({date_paiement_str}, {etablissement_payeur_str})"
            elif mode_paiement == 'Virement':
                details_paiement = f"Virement: ({date_paiement_str})"
            elif mode_paiement == 'Mobile Money':
                details_paiement = "Mobile Money"
            elif mode_paiement == 'Espèce':
                details_paiement = "Paiement en espèces"
            row_data[6] = details_paiement # Mettre à jour la colonne 'Détails Paiement'

            payment_status = row_dict.get('Etat paiement (Payée ou non)')
            fill_to_apply = None
            if payment_status == 'Payé':
                fill_to_apply = green_fill
            elif payment_status == 'Non payé':
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            ligneActuelle += 1

    ligneActuelle += 1

    # Calcul et affichage des totaux
    if data:
        df_calc = pd.DataFrame(data)

        total_by_type_paid = df_calc[df_calc['Etat paiement (Payée ou non)'] == 'Payé'].groupby('Traitement (Type)')[
            'montant_facture'].sum()

        if not total_by_type_paid.empty:
            ws.cell(row=ligneActuelle, column=1, value="Facture total pour :").font = bold_font
            ligneActuelle += 1
            for service_type, total_amount in total_by_type_paid.items():
                ws.cell(row=ligneActuelle, column=2, value=f"{service_type} (Payé)").font = bold_font
                ws.cell(row=ligneActuelle, column=3, value=total_amount).font = bold_font
                ligneActuelle += 1
        else:
            ws.cell(row=ligneActuelle, column=1,
                    value="Aucun montant payé pour les types de traitement ce mois.").font = bold_font
            ligneActuelle += 1

        ligneActuelle += 1

        # Total de paiement par mode de paiement
        ws.cell(row=ligneActuelle, column=1, value="Total de paiement par mode de paiement :").font = bold_font
        ligneActuelle += 1
        payment_mode_counts = df_calc.groupby('Mode de Paiement').size().reset_index(name='Nombre de Paiements')
        for _, row in payment_mode_counts.iterrows():
            ws.cell(row=ligneActuelle, column=2, value=f"{row['Mode de Paiement']} :").font = bold_font
            ws.cell(row=ligneActuelle, column=3, value=row['Nombre de Paiements']).font = bold_font
            ligneActuelle += 1
        ligneActuelle += 1

        grand_total = df_calc['montant_facture'].sum()
        ws.cell(row=ligneActuelle, column=1, value="Montant total des traitements effectués ce mois :").font = bold_font
        ws.cell(row=ligneActuelle, column=3, value=grand_total).font = bold_font
        ligneActuelle += 1

    max_col_for_width = len(table_headers)

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                # Gérer les dates pour le calcul de la largeur
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        # Save to a file in the current directory
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel de la facture : {e}")


# --- Fonction pour générer le fichier Excel du rapport de traitement client complet (annuel/complet) ---
def generate_comprehensive_treatment_report_excel(data: list[dict], client_full_name: str, report_period: str):
    """
    Génère un fichier Excel pour un rapport de traitement complet (annuel ou sur une période donnée)
    pour un client.
    """
    safe_client_name = "".join(c for c in client_full_name if c.isalnum() or c in (' ', '-', '_')).replace(' ',
                                                                                                           '_').rstrip(
        '_')
    file_name = f"Rapport_Traitements_{safe_client_name}_{report_period.replace(' ', '_')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Traitements {client_full_name} {report_period}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage pour les statuts de planning
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair (Effectué)
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Jaune clair (À venir)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair (Annulé/Reporté)

    current_row = 1

    # Informations du client (en-tête)
    if data:
        client_info = data[0]
        client_display_name = f"{client_info['client_nom']} {client_info['client_prenom']}"
        if client_info['client_categorie'] != 'Particulier':
            client_display_name = f"{client_info['client_nom']} (Responsable: {client_info['client_prenom'] if client_info['client_prenom'] else 'N/A'})"

        ws.cell(row=current_row, column=1, value="Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_display_name)
        current_row += 1

        # Ajout du numéro de contrat
        ws.cell(row=current_row, column=1, value="N° Contrat :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info.get('Référence Contrat', 'N/A'))
        current_row += 1

        ws.cell(row=current_row, column=1, value="Adresse :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_adresse'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Téléphone :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_telephone'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Catégorie Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_categorie'])
        current_row += 1

        ws.cell(row=current_row, column=1, value="Axe Client :").font = bold_font
        ws.cell(row=current_row, column=2, value=client_info['client_axe'])
        current_row += 1

    current_row += 1

    # Ligne de titre du rapport
    ws.cell(row=current_row, column=1,
            value=f"Rapport de Traitement pour la période : {report_period}").font = header_font
    table_headers = [
        'Référence Contrat', 'Type de Traitement', 'Redondance (Mois)',
        'Date de Planification', 'Statut du Planning', 'Remarque',
        'Motif Signalement', 'Type Signalement'
    ]
    max_cols_for_merge = len(table_headers)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols_for_merge)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    current_row += 2

    # Tableau des détails de traitement
    # Écrire les en-têtes du tableau
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
    current_row += 1

    if not data:
        ws.cell(row=current_row, column=1,
                value=f"Aucun traitement trouvé pour le client '{client_full_name}' pour la période sélectionnée.").border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(table_headers))
        current_row += 1
    else:
        df_treatment_data = pd.DataFrame(data)

        for r_idx, row_dict in enumerate(df_treatment_data.to_dict('records'), start=current_row):
            row_data = [
                row_dict.get('Référence Contrat', 'N/A'),
                row_dict.get('Type de Traitement', 'N/A'),
                row_dict.get('Redondance (Mois)', 'N/A'),
                row_dict.get('Date de Planification', 'N/A'),
                row_dict.get('Statut du Planning', 'N/A'),
                row_dict.get('Remarque', 'N/A'),
                row_dict.get('Motif Signalement', 'N/A'),
                row_dict.get('Type Signalement', 'N/A')
            ]

            planning_status = row_dict.get('Statut du Planning')
            fill_to_apply = None
            if planning_status == 'Effectué':
                fill_to_apply = green_fill
            elif planning_status == 'À venir':
                fill_to_apply = yellow_fill
            elif planning_status in ['Annulé', 'Reporté']: # Ajoutez d'autres statuts si nécessaire
                fill_to_apply = red_fill

            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if fill_to_apply:
                    cell.fill = fill_to_apply
            current_row += 1

    current_row += 1

    # Synthèse des traitements
    if data:
        df_calc = pd.DataFrame(data)

        ws.cell(row=current_row, column=1, value="Synthèse des Traitements :").font = bold_font
        current_row += 1

        # Nombre total de traitements
        ws.cell(row=current_row, column=2, value="Nombre total de traitements :").font = bold_font
        ws.cell(row=current_row, column=3, value=len(df_calc)).font = bold_font
        current_row += 1

        # Nombre de traitements par type
        if 'Type de Traitement' in df_calc.columns:
            ws.cell(row=current_row, column=1, value="Nombre de traitements par type :").font = bold_font
            current_row += 1
            treatments_by_type = df_calc.groupby('Type de Traitement').size().reset_index(name='Count')
            for _, row in treatments_by_type.iterrows():
                ws.cell(row=current_row, column=2, value=row['Type de Traitement']).font = bold_font
                ws.cell(row=current_row, column=3, value=row['Count']).font = bold_font
                current_row += 1
            current_row += 1

        # Nombre de traitements par statut de planning
        if 'Statut du Planning' in df_calc.columns:
            ws.cell(row=current_row, column=1, value="Nombre de traitements par statut :").font = bold_font
            current_row += 1
            treatments_by_status = df_calc.groupby('Statut du Planning').size().reset_index(name='Count')
            for _, row in treatments_by_status.iterrows():
                ws.cell(row=current_row, column=2, value=row['Statut du Planning']).font = bold_font
                ws.cell(row=current_row, column=3, value=row['Count']).font = bold_font
                current_row += 1
            current_row += 1

        # Nombre de remarques et signalements
        num_remarks = df_calc['Remarque'].notna().sum()
        num_signals = df_calc['Motif Signalement'].notna().sum()
        ws.cell(row=current_row, column=1, value="Nombre de remarques :").font = bold_font
        ws.cell(row=current_row, column=2, value=num_remarks).font = bold_font
        current_row += 1
        ws.cell(row=current_row, column=1, value="Nombre de signalements :").font = bold_font
        ws.cell(row=current_row, column=2, value=num_signals).font = bold_font
        current_row += 1


    max_col_for_width = len(table_headers)

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                # Gérer les dates pour le calcul de la largeur
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        # Save to a file in the current directory
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel du rapport de traitement : {e}")

# --- Fonction de récupération des traitements pour un mois donné (pour tous les clients) ---
async def get_traitements_for_month(pool, year: int, month: int):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT co.reference_contrat   AS `Référence Contrat`, -- Ajout du numéro de contrat
                           pd.date_planification  AS `Date du traitement`,
                           tt.typeTraitement      AS `Traitement concerné`,
                           tt.categorieTraitement AS `Catégorie du traitement`,
                           CONCAT(c.nom, ' ', COALESCE(c.prenom, '')) AS `Client concerné`,
                           c.categorie            AS `Catégorie du client`,
                           c.axe                  AS `Axe du client`,
                           pd.statut              AS `Etat traitement`
                    FROM PlanningDetails pd
                             JOIN Planning p ON pd.planning_id = p.planning_id
                             JOIN Traitement t ON p.traitement_id = t.traitement_id
                             JOIN TypeTraitement tt ON t.id_type_traitement = tt.id_type_traitement
                             JOIN Contrat co ON t.contrat_id = co.contrat_id
                             JOIN Client c ON co.client_id = c.client_id
                    WHERE YEAR(pd.date_planification) = %s
                      AND MONTH(pd.date_planification) = %s
                    ORDER BY pd.date_planification;
                    """
            await cursor.execute(query, (year, month))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des traitements : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction de récupération de toutes les combinaisons année-mois contenant des traitements ---
async def get_all_existing_treatment_months(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT DISTINCT
                        YEAR(date_planification) AS annee,
                        MONTH(date_planification) AS mois
                    FROM PlanningDetails
                    ORDER BY annee DESC, mois DESC;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des mois avec traitements : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction pour générer le fichier Excel des traitements (pour tous les clients) ---
def generate_traitements_excel(data: list[dict], year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()
    file_name = f"traitements-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Traitements {month_name_fr} {year}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Définition des couleurs de remplissage
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rouge clair
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Vert clair

    # Titre du rapport
    ws.cell(row=1, column=1, value=f"Rapport des Traitements du mois de {month_name_fr} {year}").font = header_font
    ws.cell(row=1, column=1).alignment = center_align
    # Ajuster le nombre de colonnes fusionnées au nombre réel de colonnes de données
    num_data_cols = len(data[0]) if data else 7 # Fallback if data is empty
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_data_cols)

    # Nombre total de traitements
    total_traitements = len(data)
    ws.cell(row=3, column=1, value=f"Nombre total de traitements ce mois-ci : {total_traitements}").font = bold_font

    # Ligne vide pour la séparation
    ws.cell(row=4, column=1, value="")

    df = pd.DataFrame(data)

    if df.empty:
        ws.cell(row=5, column=1, value="Aucun traitement trouvé pour ce mois.").border = thin_border
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=num_data_cols)
    else:
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = bold_font
            cell.border = thin_border # Appliquer la bordure aux en-têtes

        # Itérer sur les données et appliquer la couleur
        for r_idx, row_dict in enumerate(data, start=6):
            for c_idx, col_name in enumerate(headers, 1): # Itérer sur les noms de colonnes pour maintenir l'ordre
                value = row_dict.get(col_name, 'N/A') # Obtenir la valeur par nom de colonne
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border # Appliquer la bordure aux cellules de données

                # Appliquer la couleur si c'est la colonne 'Etat traitement'
                if col_name == 'Etat traitement':
                    if value == 'Effectué':
                        cell.fill = green_fill # CORRIGÉ: vert pour "Effectué"
                    elif value == 'À venir':
                        cell.fill = yellow_fill # AJOUTÉ: jaune pour "À venir"
                    elif value in ['Annulé', 'Reporté']: # AJOUTÉ: rouge pour "Annulé" ou "Reporté"
                        cell.fill = red_fill

    max_col_for_width = len(df.columns) if not df.empty else num_data_cols

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_length = len(cell.value.strftime('%Y-%m-%d'))
                else:
                    cell_length = len(str(cell.value))
                length = max(length, cell_length)
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel des traitements : {e}")

# --- Fonction principale pour exécuter la génération de rapport ---
async def main_report_menu():
    pool = None
    try:
        print("\n--- Configuration de la connexion à la base de données ---")
        db_host = input("Entrez l'hôte de la base de données (ex: localhost): ")
        db_user = input("Entrez le nom d'utilisateur de la base de données (ex: root): ")
        db_password = input("Entrez le mot de passe de la base de données: ")
        db_name = input("Entrez le nom de la base de données (ex: Planificator): ")

        pool = await DBConnection(db_host, db_user, db_password, db_name)
        if pool is None:
            print("Échec de la connexion à la base de données. Annulation de l'opération.")
            return

        continue_generating_reports = True
        while continue_generating_reports:
            print("\n--- Menu de Génération de Rapports ---")
            print("1. Générer un rapport de tous les traitements pour un mois précis (tous clients)")
            print("2. Générer un rapport de toutes les factures pour un client sur une année spécifique")
            print("3. Quitter")

            choice = input("Entrez votre choix (1-3): ")

            if choice == '1':
                # Rapport de tous les traitements pour un mois précis (tous clients)
                all_existing_months_data = await get_all_existing_treatment_months(pool)

                selected_year = None
                selected_month = None

                if all_existing_months_data:
                    distinct_years = sorted(list(set(entry['annee'] for entry in all_existing_months_data)), reverse=True)

                    print("\nAnnées contenant des traitements déjà enregistrés :")
                    for i, year in enumerate(distinct_years):
                        print(f"  {i + 1}. {year}")
                    print("  0. Entrer une autre année manuellement")

                    while True:
                        try:
                            choice_year = int(input("Choisissez un numéro d'année dans la liste ou '0' pour entrer manuellement : "))
                            if 0 < choice_year <= len(distinct_years):
                                selected_year = distinct_years[choice_year - 1]
                                break
                            elif choice_year == 0:
                                while True:
                                    try:
                                        year_input = input("Veuillez entrer l'année pour le rapport (ex: 2023) : ")
                                        selected_year = int(year_input)
                                        if not (2000 <= selected_year <= datetime.datetime.now().year + 5):
                                            print(f"Année invalide. Veuillez entrer une année entre 2000 et {datetime.datetime.now().year + 5}.")
                                            continue
                                        break
                                    except ValueError:
                                        print("Entrée invalide. Veuillez entrer un nombre pour l'année.")
                                break
                            else:
                                print("Choix invalide. Veuillez réessayer.")
                        except ValueError:
                            print("Entrée invalide. Veuillez entrer un numéro.")

                    if selected_year:
                        months_for_selected_year = sorted(list(set(entry['mois'] for entry in all_existing_months_data if entry['annee'] == selected_year)), reverse=True)

                        if months_for_selected_year:
                            print(f"\nMois disponibles pour l'année {selected_year} :")
                            for i, month_num in enumerate(months_for_selected_year):
                                month_name = datetime.date(selected_year, month_num, 1).strftime('%B').capitalize()
                                print(f"  {i + 1}. {month_name} ({month_num})")
                            print("  0. Entrer un autre mois manuellement")

                            while True:
                                try:
                                    choice_month = int(input(f"Choisissez un numéro de mois pour {selected_year} ou '0' pour entrer manuellement : "))
                                    if 0 < choice_month <= len(months_for_selected_year):
                                        selected_month = months_for_selected_year[choice_month - 1]
                                        break
                                    elif choice_month == 0:
                                        while True:
                                            try:
                                                month_input = input("Veuillez entrer le numéro du mois (1-12) : ")
                                                selected_month = int(month_input)
                                                if not (1 <= selected_month <= 12):
                                                    print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                                                    continue
                                                break
                                            except ValueError:
                                                print("Entrée invalide. Veuillez entrer un nombre pour le mois.")
                                        break
                                    else:
                                        print("Choix invalide. Veuillez réessayer.")
                                except ValueError:
                                    print("Entrée invalide. Veuillez entrer un numéro.")
                        else:
                            print(f"\nAucun traitement trouvé pour l'année {selected_year}. Veuillez entrer le mois manuellement.")
                            while True:
                                try:
                                    month_input = input("Veuillez entrer le numéro du mois (1-12) pour le rapport (ex: 6 pour Juin) : ")
                                    selected_month = int(month_input)
                                    if not (1 <= selected_month <= 12):
                                        print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                                        continue
                                    break
                                except ValueError:
                                    print("Entrée invalide. Veuillez entrer un nombre pour le mois.")

                else:
                    print("\nAucun traitement trouvé dans la base de données. Veuillez entrer le mois et l'année manuellement.")
                    while True:
                        try:
                            year_input = input("Veuillez entrer l'année pour le rapport (ex: 2023) : ")
                            month_input = input("Veuillez entrer le numéro du mois (1-12) pour le rapport (ex: 6 pour Juin) : ")

                            selected_year = int(year_input)
                            selected_month = int(month_input)

                            if not (1 <= selected_month <= 12):
                                print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                                continue
                            if not (2000 <= selected_year <= datetime.datetime.now().year + 5):
                                print(f"Année invalide. Veuillez entrer une année entre 2000 et {datetime.datetime.now().year + 5}.")
                                continue
                            break
                        except ValueError:
                            print("Entrée invalide. Veuillez entrer un nombre pour l'année et le mois.")

                if selected_year is None or selected_month is None:
                    print("Sélection de l'année ou du mois annulée. Retour au menu principal.")
                    continue # Retourner au menu principal

                print(f"\nPréparation du rapport des traitements pour {datetime.date(selected_year, selected_month, 1).strftime('%B').capitalize()} {selected_year}...")
                traitements_data = await get_traitements_for_month(pool, selected_year, selected_month)
                generate_traitements_excel(traitements_data, selected_year, selected_month)

            elif choice == '2':
                # Rapport de toutes les factures pour un client sur une année spécifique
                print("\n--- Aperçu des clients et de leurs factures ---")
                client_infos = await obtenirInformationsClients(pool)
                if not client_infos:
                    print("Aucun client trouvé dans la base de données.")
                    continue

                print("\nListe des clients et leurs informations de facture:")
                for i, client in enumerate(client_infos):
                    print(f"{i + 1}. {client['nomComplet']} (Factures: {client['total_factures']}, Mois: {client['facture_months'] if client['facture_months'] else 'Aucun'})")

                client_index = input("Entrez le numéro du client pour le rapport annuel: ")
                try:
                    client_index = int(client_index) - 1
                    if 0 <= client_index < len(client_infos):
                        selected_client = client_infos[client_index]
                        client_id = selected_client['client_id']
                        client_full_name = selected_client['nomComplet']

                        year_input = int(input("Entrez l'année pour le rapport annuel (ex: 2023): "))
                        start_date = datetime.date(year_input, 1, 1)
                        end_date = datetime.date(year_input, 12, 31)
                        report_period_str = f"Année {year_input}"

                        print(f"Récupération des données complètes pour {client_full_name} ({report_period_str})...")
                        # Utilisez la fonction correcte pour les traitements ici
                        comprehensive_data = await get_treatments_data_for_client_comprehensive(pool, client_id, start_date, end_date)
                        if comprehensive_data:
                            print(f"Génération du rapport complet pour {client_full_name}...")
                            # Utilisez la fonction correcte pour générer le rapport de traitements
                            generate_comprehensive_treatment_report_excel(comprehensive_data, client_full_name, report_period_str)
                        else:
                            print(f"Aucune donnée de traitement complète trouvée pour {client_full_name} pour l'année {year_input}.")
                    else:
                        print("Numéro de client invalide.")
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un nombre.")
                except Exception as e:
                    print(f"Une erreur inattendue est survenue : {e}")

            elif choice == '3':
                print("Quitter le programme.")
                continue_generating_reports = False
            else:
                print("Choix invalide. Veuillez réessayer.")

            if continue_generating_reports: # Demander si l'utilisateur veut générer un autre fichier
                while True:
                    reponse = input("\nVoulez-vous générer un autre rapport ? (oui/non) : ").lower().strip()
                    if reponse in ['oui', 'o']:
                        break # Continuer la boucle principale
                    elif reponse in ['non', 'n']:
                        print("Fin de la génération des rapports.")
                        return # Sortir de la fonction main_report_menu
                    else:
                        print("Réponse invalide. Veuillez répondre par 'oui' ou 'non'.")


    except Exception as e:
        print(f"Une erreur critique est survenue dans le menu principal : {e}")
    finally:
        if pool:
            pool.close()
            print("Connexion à la base de données fermée.")

# Pour exécuter le menu principal
if __name__ == "__main__":
    # Ensure the directory for generated files exists
    if not os.path.exists("./generated_reports"): # Changed folder name for clarity
        os.makedirs("./generated_reports")
    os.chdir("./generated_reports") # Change to this directory to save files there

    asyncio.run(main_report_menu())
