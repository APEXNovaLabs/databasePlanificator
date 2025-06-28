import pandas as pd
import datetime
import asyncio
import aiomysql
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill # Importez PatternFill ici
from openpyxl.utils import get_column_letter
from Contrat.fonctionnalités.connexionDB import DBConnection

# --- Fonction de récupération des traitements pour un mois donné ---
async def obtenirTraitementParMois(pool, year: int, month: int):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT pd.date_planification        AS `Date du traitement`,
                           tt.typeTraitement            AS `Traitement concerné`,
                           tt.categorieTraitement       AS `Catégorie du traitement`,
                           CONCAT(c.nom, ' ', c.prenom) AS `Client concerné`,
                           c.categorie                  AS `Catégorie du client`,
                           c.axe                        AS `Axe du client`,
                           pd.statut                    AS `Etat traitement` -- AJOUT DE CETTE COLONNE POUR RÉCUPÉRER LE STATUT
                    FROM PlanningDetails pd
                             JOIN
                         Planning p ON pd.planning_id = p.planning_id
                             JOIN
                         Traitement t ON p.traitement_id = t.traitement_id
                             JOIN
                         TypeTraitement tt ON t.id_type_traitement = tt.id_type_traitement
                             JOIN
                         Contrat co ON t.contrat_id = co.contrat_id
                             JOIN
                         Client c ON co.client_id = c.client_id
                    WHERE YEAR(pd.date_planification) = %s
                      AND MONTH(pd.date_planification) = %s
                    ORDER BY pd.date_planification;
                    """
            await cursor.execute(query, (year, month))
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des traitements : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction de récupération de toutes les combinaisons année-mois contenant des traitements ---
async def obtenirToutTraitementDuMois(pool):
    conn = None
    try:
        conn = await pool.acquire()
        async with conn.cursor(aiomysql.DictCursor) as cursor:
            query = """
                    SELECT DISTINCT
                        YEAR(date_planification) AS annee,
                        MONTH(date_planification) AS mois
                    FROM PlanningDetails
                    ORDER BY annee DESC, mois DESC;
                    """
            await cursor.execute(query)
            result = await cursor.fetchall()
            return result
    except Exception as e:
        print(f"Erreur lors de la récupération des mois avec traitements : {e}")
        return []
    finally:
        if conn:
            pool.release(conn)

# --- Fonction pour générer le fichier Excel des traitements ---
def generationTraitementExcel(data: list[dict], year: int, month: int):
    month_name_fr = datetime.date(year, month, 1).strftime('%B').capitalize()
    file_name = f"traitements-{month_name_fr}-{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Traitements {month_name_fr} {year}"

    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')

    # Définition des couleurs de remplissage
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rouge clair
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Vert clair

    # Titre du rapport
    ws.cell(row=1, column=1, value=f"Rapport des Traitements du mois de {month_name_fr} {year}").font = header_font
    ws.cell(row=1, column=1).alignment = center_align
    num_data_cols = 7
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_data_cols)

    # Nombre total de traitements
    total_traitements = len(data)
    ws.cell(row=3, column=1, value=f"Nombre total de traitements ce mois-ci : {total_traitements}").font = bold_font

    # Ligne vide pour la séparation
    ws.cell(row=4, column=1, value="")

    df = pd.DataFrame(data)

    if df.empty:
        ws.cell(row=5, column=1, value="Aucun traitement trouvé pour ce mois.")
    else:
        headers = df.columns.tolist()
        status_col_index = -1
        try:
            status_col_index = headers.index('Etat traitement') # Trouvez l'index de la colonne de statut
        except ValueError:
            print("AVERTISSEMENT: La colonne 'Etat traitement' n'a pas été trouvée dans les données. Les couleurs ne seront pas appliquées.")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = bold_font

        for r_idx, row_dict in enumerate(data, start=6):
            for c_idx, col_name in enumerate(headers, 1):
                value = row_dict.get(col_name)
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Appliquer la couleur si c'est la colonne 'Etat traitement' et si la valeur est connue
                if col_name == 'Etat traitement':
                    if value == 'Effectué':
                        cell.fill = red_fill
                    elif value == 'À venir':
                        cell.fill = green_fill
                # else:
                #     cell.fill = PatternFill(fill_type=None) # Optionnel: Réinitialiser la couleur pour les autres colonnes

    max_col_for_width = len(df.columns) if not df.empty else num_data_cols

    for i in range(1, max_col_for_width + 1):
        column_letter = get_column_letter(i)
        length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = length + 2

    try:
        output = BytesIO()
        wb.save(output)
        with open(file_name, 'wb') as f:
            f.write(output.getvalue())

        print(f"Fichier '{file_name}' généré avec succès.")
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel des traitements : {e}")

# --- Fonction principale pour exécuter le rapport des traitements ---
async def generationRapportMain():
    pool = None
    try:
        pool = await DBConnection()
        if not pool:
            print("Échec de la connexion à la base de données. Annulation de l'opération.")
            return

        donnéesMoisExistant = await obtenirToutTraitementDuMois(pool)

        annéeChoisie = None
        moisChoisie = None

        if donnéesMoisExistant:
            # 1. Obtenir les années distinctes
            distinct_years = sorted(list(set(entry['annee'] for entry in donnéesMoisExistant)), reverse=True)

            print("\nAnnées contenant des traitements déjà enregistrés :")
            for i, year in enumerate(distinct_years):
                print(f"  {i + 1}. {year}")
            print("  0. Entrer une autre année manuellement")

            chosen_year_index = -1
            while True:
                try:
                    choice = int(input("Choisissez un numéro d'année dans la liste ou '0' pour entrer manuellement : "))
                    if 0 < choice <= len(distinct_years):
                        annéeChoisie = distinct_years[choice - 1]
                        chosen_year_index = choice - 1
                        break
                    elif choice == 0:
                        while True:
                            try:
                                inputAnnée = input("Veuillez entrer l'année pour le rapport (ex: 2023) : ")
                                annéeChoisie = int(inputAnnée)
                                if not (2000 <= annéeChoisie <= datetime.datetime.now().year + 5):
                                    print(f"Année invalide. Veuillez entrer une année entre 2000 et {datetime.datetime.now().year + 5}.")
                                    continue
                                break
                            except ValueError:
                                print("Entrée invalide. Veuillez entrer un nombre pour l'année.")
                        break
                    else:
                        print("Choix invalide. Veuillez réessayer.")
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un numéro.")

            # 2. Après avoir choisi l'année, filtrer et afficher les mois disponibles pour cette année
            if annéeChoisie:
                months_for_selected_year = sorted(list(set(entry['mois'] for entry in donnéesMoisExistant if entry['annee'] == annéeChoisie)), reverse=True)

                if months_for_selected_year:
                    print(f"\nMois disponibles pour l'année {annéeChoisie} :")
                    for i, month_num in enumerate(months_for_selected_year):
                        month_name = datetime.date(annéeChoisie, month_num, 1).strftime('%B').capitalize()
                        print(f"  {i + 1}. {month_name} ({month_num})")
                    print("  0. Entrer un autre mois manuellement")

                    while True:
                        try:
                            choice = int(input(f"Choisissez un numéro de mois pour {annéeChoisie} ou '0' pour entrer manuellement : "))
                            if 0 < choice <= len(months_for_selected_year):
                                moisChoisie = months_for_selected_year[choice - 1]
                                break
                            elif choice == 0:
                                while True:
                                    try:
                                        month_input = input("Veuillez entrer le numéro du mois (1-12) : ")
                                        moisChoisie = int(month_input)
                                        if not (1 <= moisChoisie <= 12):
                                            print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                                            continue
                                        break
                                    except ValueError:
                                        print("Entrée invalide. Veuillez entrer un nombre pour le mois.")
                                break
                            else:
                                print("Choix invalide. Veuillez réessayer.")
                        except ValueError:
                            print("Entrée invalide. Veuillez entrer un numéro.")
                else: # Entrée manuelle de l'année
                    print(f"\nAucun traitement trouvé pour l'année {annéeChoisie}. Veuillez entrer le mois manuellement.")
                    while True:
                        try:
                            month_input = input("Veuillez entrer le numéro du mois (1-12) pour le rapport (ex: 6 pour Juin) : ")
                            moisChoisie = int(month_input)
                            if not (1 <= moisChoisie <= 12):
                                print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                                continue
                            break
                        except ValueError:
                            print("Entrée invalide. Veuillez entrer un nombre pour le mois.")

        else: # Aucune données dans la BD
            print("\nAucun traitement trouvé dans la base de données. Veuillez entrer le mois et l'année manuellement.")
            while True:
                try:
                    inputAnnée = input("Veuillez entrer l'année pour le rapport (ex: 2023) : ")
                    month_input = input("Veuillez entrer le numéro du mois (1-12) pour le rapport (ex: 6 pour Juin) : ")

                    annéeChoisie = int(inputAnnée)
                    moisChoisie = int(month_input)

                    if not (1 <= moisChoisie <= 12):
                        print("Numéro de mois invalide. Veuillez entrer un nombre entre 1 et 12.")
                        continue
                    if not (2000 <= annéeChoisie <= datetime.datetime.now().year + 5):
                        print(f"Année invalide. Veuillez entrer une année entre 2000 et {datetime.datetime.now().year + 5}.")
                        continue
                    break
                except ValueError:
                    print("Entrée invalide. Veuillez entrer un nombre pour l'année et le mois.")

        if annéeChoisie is None or moisChoisie is None:
            print("Sélection de l'année ou du mois annulée. Fin du rapport.")
            return

        print(
            f"\nPréparation du rapport des traitements pour {datetime.date(annéeChoisie, moisChoisie, 1).strftime('%B').capitalize()} {annéeChoisie}...")
        traitements_data = await obtenirTraitementParMois(pool, annéeChoisie, moisChoisie)
        generationTraitementExcel(traitements_data, annéeChoisie, moisChoisie)

    except Exception as e:
        print(f"Une erreur inattendue est survenue dans le script principal : {e}")
    finally:
        if pool:
            await pool.close()

if __name__ == "__main__":
    asyncio.run(generationRapportMain())